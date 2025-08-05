from openpyxl import load_workbook
from settings import PATH_EXCEL


class ExcelData:
    def __init__(self):
        self._workbook = None

    def get_workbook(self, name_workbook: str):
        if self._workbook is None:
            self._workbook = load_workbook(PATH_EXCEL, data_only=True)
        return self._workbook[name_workbook]

    def get_order_data(self):
        sheet = self.get_workbook(name_workbook="declension")

        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[75]:
                break
            data[row[75].date()] = row[76]
        
        return data

    def get_vacation_data(self):
        sheet = self.get_workbook(name_workbook="ВІДПУ")

        data = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row[1]:
                break
            entry = {
                "ПІБ": row[1],
                "Підрозділ": row[2],
                "Тип відпустки": row[3],
                "Запланована дата прибуття": row[21],
                "Фактична дата прибуття": row[22]
            }
            data.append(entry)
        return data

    def get_sh(self):
        sheet = self.get_workbook(name_workbook="sh")

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[2]:
                break
            entry = {
                "Повна посада": row[1],
                "звання": row[2],
                "ПІБ": row[3],
            }
            data.append(entry)
        return data

